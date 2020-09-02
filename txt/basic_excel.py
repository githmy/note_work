import os
import openpyxl

from openpyxl.drawing.image import Image

def main():
    # 1. 默认加载原数据
    # inpath = os.path.join("..", "aa.xls")
    filename = os.path.join("..", "a2.xlsx")
    workbook = openpyxl.load_workbook(filename)
    # worksheet = workbook.get_sheet_by_name("Sheet1")  # 读取Sheet
    worksheet = workbook.worksheets[0]  # 读取Sheet
    rows, cols = worksheet.max_row, worksheet.max_column
    print(rows, cols)
    print(worksheet._images)
    yellow = []
    red = []
    # for i in range(1, rows):
    #     for j in range(1, cols):
    for i in range(rows):
        for j in range(cols):
            ce = worksheet.cell(row=i + 1, column=j + 1)
            fill = ce.fill
            font = ce.font
            # print(font.color)
            # print(font.color.rgb)
            # print(fill.fgColor.rgb)
            # print(fill.start_color.rgb)  # 单元颜色
            print(ce.value)  # 单元值
            if fill.start_color.rgb == "FFFFFF00" and ce.value != None:
                yellow.append(ce.value)
            if font.color.rgb == "FFFF0000":
                red.append(ce.value)
    print(99999)
    print(yellow, red)
    pass


if __name__ == '__main__':
    main()
